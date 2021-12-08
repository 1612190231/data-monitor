# -*- coding: utf-8 -*-
# Description: excel操作类
# Created: luchengkai 2021/08/013

import csv
import openpyxl


def read_value_by_sheet(read_path, read_sheet):
    """
    获取表单中的所有数据
    :param read_path:
    :param read_sheet:
    :return:
    """
    wb = openpyxl.load_workbook(read_path)
    # 选中文档中的一个datasheet
    ws = wb[read_sheet]
    return list(ws.values)


def insert_to_excel(title_list, insert_list, insert_path, sheet="sheet1"):
    """
    空表写入，定义空表写入excel函数
    :param title_list:  List[String]
    :param insert_list: List[Tuple()]
    :param insert_path: String
    :param sheet: String
    """
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet(sheet, index=0)

    # 写入表头
    column = 1
    for t in title_list:
        sheet1.cell(1, column, t)
        column += 1

    # 写入内容
    row = 2
    for data in insert_list:  # 控制行
        col = 1
        for one in data:  # 控制每一列
            sheet1.cell(row, col, one)  # rou代表列，col代表行
            col += 1
        row += 1
    workbook.save(insert_path)
    print("xlsx格式表格【空表】写入数据成功！")


# 追加写入，定义追加写入excel函数
def append_to_excel(append_list, append_path, sheet="sheet1"):
    """
    空表写入，定义空表写入excel函数
    :param append_list: List[Tuple()]
    :param append_path: String
    :param sheet: String
    """
    workbook = openpyxl.load_workbook(append_path)  # 打开工作薄
    worksheet = workbook.get_sheet_by_name(sheet)  # 获取工作薄
    rows_exists = worksheet.max_row  # 获取已经存在的数据行数
    # new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    # new_worksheet = new_workbook.get_sheet(sheet)  # 获取转化后的第一个表格
    # v_list = []
    for data in append_list:
        # v_list.append(v)
        # for j in range(0, len(v_list)):
        #     worksheet.write(rows_exists, j, v_list[j])
        col = 1
        for one in data:  # 控制每一列
            worksheet.cell(rows_exists + 1, col, one)  # rou代表列，col代表行
            col += 1
        rows_exists += 1
    workbook.save(append_path)  # 保存工作簿
    print("xlsx格式表格【追加】写入数据成功！")


if __name__ == '__main__':
    print(1)
